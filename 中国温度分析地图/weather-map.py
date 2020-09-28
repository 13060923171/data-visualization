from pyecharts import options as opts
from pyecharts.charts import Geo
from pyecharts.faker import Faker
from pyecharts.globals import ChartType,CurrentConfig,NotebookType
CurrentConfig.NOTEBOOK_TYPE = NotebookType.JUPYTER_LAB
from openpyxl import load_workbook
import requests

#定义一个字典
city_code = {}
#用load_workbook方法来打开这个xlsx文件，并且返回这个xlsx文件的内容
sheet = load_workbook('AMap_adcode_citycode.xlsx').active
#去定位这个表格的所有行
for row in sheet.rows:
    #定义这个row的值，city就是row的A列的内容，code就是row的B列所有内容，_就是row的c列的所有内容
    city,code, _ = row
    #再给city赋值
    city = city.value
    #给code赋值
    code = code.value
    try:
        #写一个判断语句当city里面出现了市或者自治区这几个字的时候在list列表里面添加city的值
        if city[-1] == '市' or city[-3] == '自治区':
            list.append(city)
            #用字典的形式，给每个city加上对应的code
            city_code[city] = code
    except:
        pass

city_weather = []
#items( )用于 返回一个字典的拷贝列表，这句话的意思就是city_code.items()变成一个可以打印的字典，city就是这个字典的key，code就是这个字典的value
for city,code in city_code.items():
    #请求这个API接口
    req = requests.get('https://restapi.amap.com/v3/weather/weatherInfo?key=e151ce35ee68dca653fa2a301b9f5050&city={}'.format(code))
    #返回一个json的格式
    result = req.json()
    #再去定位这个json的位置
    city_weather.append([city,result['lives'][0]['temperature']])
#进行数据可视化操作
c = (
    #是否忽视不存在的坐标，默认值为false，这里用TRUE的话就是忽视的意思
    Geo(is_ignore_nonexistent_coord=True)
    #定义这个地图类型
    .add_schema(maptype="china")
    .add(
        #系列名称，就是你要用什么类型的可视化就用这个可视化类型的名称
        "geo",
        #数据项（坐标点名称，坐标点值）
        city_weather,
        # GeoType.GeoType.EFFECT_SCATTER，GeoType.HEATMAP，GeoType.LINES
        type_=ChartType.HEATMAP,
        #标签配置项
        label_opts=opts.LabelOpts(is_show=False)
    )
    .set_global_opts(
        visualmap_opts=opts.VisualMapOpts(),
        #确定这个图表的标题
        title_opts=opts.TitleOpts(title="Geo-HeatMap"),
    )
    #生成一个名为weather_map.html动态交互网页
    .render('weather_map.html')
)
